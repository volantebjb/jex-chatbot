import os
import json
import string
from openpyxl import Workbook
from deep_translator import GoogleTranslator
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.sentiment.vader import SentimentIntensityAnalyzer

if __name__ == '__main__':
    
    # Data Export from Firestore Database 
    print("Data are currently being extracted from Firestore Database.")
    print("Please wait...")
    
    # Export Firestore Database to TXT file
    firestoreDatabase = os.system(
        "firestore-export -a FirestoreServiceAccount.json -b JexChatbotDatabase.txt -p"
        )
    
    # Data Export from Firestore Database Successful. 
    print("Data were successfully extracted from Firestore Database.")
    
    # Import TXT file
    json_data = {}
    
    with open("JexChatbotDatabase.txt") as json_file:
       json_data = json.load(json_file)

    # Create Excel file
    workbook = Workbook()
    worksheet =  workbook.active
    worksheet.title = "2021-06"
    
    worksheet.cell(1, 1, "Conversation ID")
    worksheet.cell(1, 2, "Courier     ")
    worksheet.cell(1, 3, "Location    ")
    worksheet.cell(1, 4, "Location Type")
    worksheet.cell(1, 5, "Time    ")
    worksheet.cell(1, 6, "Flexibility ")
    worksheet.cell(1, 7, "Reliability ")
    worksheet.cell(1, 8, "Standardization")
    worksheet.cell(1, 9, "Attitude   ")
    worksheet.cell(1, 10, "Overall Feedback")
    worksheet.cell(1, 11, "Negative Score")
    worksheet.cell(1, 12, "Neutral Score")
    worksheet.cell(1, 13, "Positive Score")
    worksheet.cell(1, 14, "Compound Score")
    worksheet.cell(1, 15, "Sentiment   ")

    # Sentiment Analysis
    print("User queries are currently in sentiment analysis phase.")
    print("Please wait...")
    
    # Put data from JSON file to Excel file
    row = 1
    json_data = json_data['__collections__']['users']

    for userID in json_data.keys():
        row += 1
        
        worksheet.cell(row, 1, userID)
        worksheet.cell(row, 2, str(json_data[userID]["Courier"]))
        worksheet.cell(row, 3, str(json_data[userID]["Location"]))
        worksheet.cell(row, 4, str(json_data[userID]["LocationType"]))
        worksheet.cell(row, 5, str(json_data[userID]["Time"]))
        worksheet.cell(row, 6, str(json_data[userID]["Flexibility"]))
        worksheet.cell(row, 7, str(json_data[userID]["Reliability"]))
        worksheet.cell(row, 8, str(json_data[userID]["Standardization"]))
        worksheet.cell(row, 9, str(json_data[userID]["Attitude"]))
        worksheet.cell(row, 10, str(json_data[userID]["Feedback"]))

        # Overall Feedback Sentiment Analysis
        sentence = (json_data[userID]["Feedback"])
        
        # Translate the Query to English
        query = GoogleTranslator(source='tl', target='en').translate(sentence)
        
        # Convert the Query to Lowercase
        lower_cased_query = query.lower()
        
        # Remove punctuations in the Query
        cleaned_query  = lower_cased_query.translate(str.maketrans(
            "", "", string.punctuation))
        
        # Split the Query into Tokenized Words
        tokenized_words = word_tokenize(cleaned_query, "english")
        
        #Remove Stop Words from the Tokenized Words 
        query = []
        
        for word in tokenized_words:
            if word not in stopwords.words("english"):
                query.append(word)
        
        # Sentiment Analysis using VaderSentiment
        analyzer = SentimentIntensityAnalyzer()
        sentiment = analyzer.polarity_scores(cleaned_query)
        
        negative = sentiment["neg"]
        neutral = sentiment["neu"]
        positive = sentiment["pos"]
        compound = sentiment["compound"]
        
        # Sentiment Analysis Accuracy Score
        
        # Put Negative Polarity Score to Excel File
        worksheet.cell(row, 11, negative)
        
        # Put Negative Polarity Score to Excel File
        worksheet.cell(row, 12, neutral)
        
        # Put Negative Polarity Score to Excel File
        worksheet.cell(row, 13, positive)
        
        # Put Compound Polarity Score to Excel File
        worksheet.cell(row, 14, compound)
        
        # Put Sentiment to Excel File
        if compound < 0:
            worksheet.cell(row, 15, str("Negative"))
            
        elif compound == 0:
            worksheet.cell(row, 15, str("Neutral"))
            
        elif compound > 0:
            worksheet.cell(row, 15, str("Positive"))
        
    # Sentiment Analysis Successful
    print("Data successfully undergone sentiment analysis.")     
     
    # Fix column width in Excel file 
    dims = {}
    
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )    
                
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value
        
    # Data Import
    print("Data are currently being imported into an Excel file.")
    print("Please wait...")      
    
    # Save File 
    workbook.save("JexChatbotDatabase.xlsx")
    
    # Data Import Successful
    print("Data were successfully imported into an Excel file.")
    